from unicodedata import name
from openpyxl import Workbook
from openpyxl import load_workbook





def add_baza(id,first_name,last_name,email,gender,phone,adress,data,children,language,country):
    wb = Workbook()
    ws = wb.active
    baza = []
    book = load_workbook('Baza.xlsx')
    sheet = book.active
    for i in range(1, sheet.max_row+1):
        A = f'A{i}'
        B = f'B{i}'
        C = f'C{i}'
        D = f'D{i}'
        E = f'E{i}'
        F = f'F{i}'
        G = f'G{i}'
        H = f'H{i}'
        I = f'I{i}'
        J = f'J{i}'
        K = f'K{i}'
        baza.append([sheet[A].value, sheet[B].value, sheet[C].value, sheet[D].value, sheet[E].value, sheet[F].value, sheet[G].value, sheet[H].value, sheet[I].value, sheet[J].value, sheet[K].value])
    baza.append([id,first_name,last_name,email,gender,phone,adress,data,children,language,country])
    for i in baza:
        ws.append(i)
    wb.save("Baza.xlsx")
    return "Bazaga ma`lumot qo`shildi"







