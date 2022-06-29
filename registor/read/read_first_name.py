from openpyxl import load_workbook





def read_first_name(o):
    
    book = load_workbook('Baza.xlsx')
    sheet = book.active
  
    for i in range(2, sheet.max_row+1):
        A = f'A{i}'
        B = f'B{i}'
        C = f'C{i}'
        D = f'D{i}'
        E = f'E{i}'
        F = f'F{i}'
        G = f'G{i}'
        H = f'H{i}'
        I = f'I{i}'
        J = f'J{i}'
        K = f'K{i}'

   
    if sheet[B].value == o.title():
        print(f"\n\n\nID: {sheet[A].value} \nFirst_name: {sheet[B].value} \nLast_name:{sheet[C].value} \nEmail: {sheet[D].value} \nGender: {sheet[E].value} \nPhone: {sheet[F].value} \nAdress: {sheet[G].value} \nData: {sheet[H].value} \nChildren: {sheet[I].value} \nLanguage: {sheet[J].value} \nCountry:{sheet[K].value}\n\n\n")
    else:
        print("Bizda bu davlatdan ishchi yuq!!!")